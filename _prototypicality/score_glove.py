#!/usr/bin/env python3
# score_glove.py
# Command line tool for the GloVe path of Toubia and Netzer prototypicality scoring.
#
# Two steps:
#   build  -> read topic documents and a GloVe file, compute the vocabulary, the cosine edges,
#             and the prototype, and save them to a small model JSON. Run this once per topic.
#   score  -> load the model JSON and score one idea, or a CSV of ideas.
#
# Scoring works in two modes:
#   closed (default)   -> an idea's network is built only from its words that are in the
#                         topic VOCABULARY. Fast, needs no GloVe file, but an idea whose
#                         words fall outside that narrow vocabulary gets too few nodes and
#                         comes back not scorable (blank KPI).
#   open  (--glove ...) -> any idea that closed mode cannot score is rescued by building its
#                         network from ALL of its own content words that have a GloVe vector,
#                         while still benchmarking against the topic prototype. With this,
#                         nearly every idea with two real content words gets a KPI. Pass
#                         --open-only to score every idea this way (comparable across all rows).
#
# Dependencies: pip install nltk numpy
# Closed scoring has no GloVe dependency, so it is fast and the big file is only touched once.

import argparse, csv, json, os, sys
import numpy as np

from proto_core import (tokenize_and_stem, merge_s2w, build_vocab, glove_edges,
                        prototypical_cdf, score_stems, idea_to_text)
from glove_loader import load_glove


def read_docs(docs_path, docs_dir):
    """Read topic documents. --docs is one document per line. --docs-dir is one file per document."""
    docs = []
    if docs_dir:
        for fn in sorted(os.listdir(docs_dir)):
            p = os.path.join(docs_dir, fn)
            if os.path.isfile(p) and fn.lower().endswith((".txt", ".md")):
                with open(p, encoding="utf-8") as f:
                    docs.append(f.read())
    if docs_path:
        with open(docs_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    docs.append(line)
    return docs


def cmd_build(args):
    raw_docs = read_docs(args.docs, args.docs_dir)
    if len(raw_docs) < 2:
        sys.exit("Need at least two topic documents. Use --docs file.txt (one per line) "
                 "or --docs-dir folder/.")

    docs_stems, maps = [], []
    for d in raw_docs:
        s, m = tokenize_and_stem(d)
        docs_stems.append(s)
        maps.append(m)
    s2w = merge_s2w(maps)

    min_dc = args.min_doc_count if args.min_doc_count else max(2, round(0.02 * len(docs_stems)))
    vocab = build_vocab(docs_stems, min_dc)
    if not vocab:
        sys.exit("Empty vocabulary. Lower --min-doc-count or add more documents.")

    needed = {w for st in vocab for w in s2w.get(st, {st})}
    vectors = load_glove(args.glove, needed)
    edges, kept = glove_edges(vocab, s2w, vectors)
    if len(kept) < 2:
        sys.exit("Fewer than two vocabulary stems had a GloVe vector. Check the GloVe file path.")

    proto = prototypical_cdf(docs_stems, kept, edges)

    model = {
        "vocab": kept,
        "edges": {f"{a}|{b}": w for (a, b), w in edges.items()},
        "prototype": proto.tolist(),
        "meta": {
            "n_docs": len(raw_docs),
            "min_doc_count": min_dc,
            "vocab_size": len(vocab),
            "vocab_with_vector": len(kept),
            "glove": os.path.basename(args.glove),
        },
    }
    with open(args.model, "w", encoding="utf-8") as f:
        json.dump(model, f)

    print(f"Built model from {len(raw_docs)} documents.")
    print(f"  vocabulary stems: {len(vocab)}")
    print(f"  stems with a GloVe vector: {len(kept)}")
    print(f"  edges: {len(edges)}")
    print(f"  min_doc_count: {min_dc}")
    print(f"Saved model to {args.model}")
    if len(kept) < 40:
        print()
        print(f"NOTE: the vocabulary is small ({len(kept)} nodes), so many ideas may come back as")
        print("not scorable in closed mode. Lower the threshold, for example --min-doc-count 2 or 3,")
        print("and build again, add more on-topic lines to your documents file, or score with")
        print("--glove so out-of-vocabulary ideas are rescued in open mode.")


def load_model(path):
    with open(path, encoding="utf-8") as f:
        model = json.load(f)
    vocab = model["vocab"]
    edges = {}
    for k, w in model["edges"].items():
        a, b = k.split("|")
        edges[(a, b)] = w
    proto = np.asarray(model["prototype"], dtype=float)
    return vocab, edges, proto, model.get("meta", {})


def score_one(idea, vocab, edges, proto):
    stems, _ = tokenize_and_stem(idea_to_text(idea))
    return score_stems(stems, vocab, edges, proto)


def score_one_open(idea, vectors, proto):
    """
    Open scoring. The idea's nodes are all of its own content words that have a GloVe vector,
    not only the ones in the topic vocabulary. Edges between those words are cosines computed
    on the fly. The benchmark is still the topic prototype, so the comparison stays topic
    specific on the benchmark side while almost any idea with two content words gets a KPI.
    """
    stems, s2w = tokenize_and_stem(idea_to_text(idea))
    uniq = sorted(set(stems))
    if len(uniq) < 2:
        return {"scorable": False, "reason": "fewer than two content words",
                "n_nodes": len(uniq), "n_edges": 0}
    edges, kept = glove_edges(uniq, s2w, vectors)
    return score_stems(kept, kept, edges, proto)


def score_idea(idea, vocab, edges, proto, open_vectors, open_only):
    """
    Score a single idea and report which mode produced the score.

    Returns (result, mode) where mode is one of:
      "closed" - scored on the idea's words that are in the topic vocabulary
      "open"   - closed mode could not score it (too few in-vocabulary words), so it was
                 rescued by scoring all of its own GloVe words. Also used for every idea
                 when open_only is set.
      ""       - not scorable in any available mode (genuinely fewer than two content words).
    """
    if open_only and open_vectors is not None:
        res = score_one_open(idea, open_vectors, proto)
        return res, ("open" if res["scorable"] else "")

    res = score_one(idea, vocab, edges, proto)
    if res["scorable"]:
        return res, "closed"

    # Open-mode rescue: the idea's words fall outside the narrow topic vocabulary, so build
    # its network from all of its own content words instead. Only possible with a GloVe file.
    if open_vectors is not None:
        res_open = score_one_open(idea, open_vectors, proto)
        if res_open["scorable"]:
            return res_open, "open"
        return res_open, ""
    return res, ""


def _collect_words(ideas, tcol, dcol):
    """Gather every surface word across a list of idea rows, for one GloVe load."""
    needed = set()
    for r in ideas:
        idea = {"title": (r.get(tcol, "") if tcol else "") or "", "description": r.get(dcol, "") or ""}
        _, m = tokenize_and_stem(idea_to_text(idea))
        for words in m.values():
            needed.update(words)
    return needed


def read_ideas(path):
    """
    Read ideas from an Excel workbook (.xlsx, .xlsm) or a delimited text file
    (.csv, .tsv, .txt). Returns (rows, fieldnames) where each row is a dict.
    The first row is treated as the header.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("Reading an Excel file needs openpyxl. Install it with:\n"
                     "    pip install openpyxl\n"
                     "Or in Excel use File, Save As, CSV UTF-8, and point --ideas at the .csv.")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            return [], []
        fieldnames = [str(h).strip() if h is not None else "" for h in header]
        rows = []
        for r in it:
            if r is None or all(c is None or str(c).strip() == "" for c in r):
                continue
            row = {}
            for i, fn in enumerate(fieldnames):
                val = r[i] if i < len(r) else None
                row[fn] = "" if val is None else str(val)
            rows.append(row)
        return rows, fieldnames

    # delimited text: sniff comma, semicolon, or tab (handles European CSVs too)
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            delim = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
        except Exception:
            delim = ","
        reader = csv.DictReader(f, delimiter=delim, restkey="_overflow")
        if not reader.fieldnames:
            return [], []
        fieldnames = list(reader.fieldnames)
        rows = []
        for r in reader:
            overflow = r.pop("_overflow", None)
            if overflow:  # stray delimiters: append the pieces back to the last column
                last = fieldnames[-1]
                r[last] = (r.get(last, "") or "") + delim + delim.join(overflow)
            rows.append(dict(r))
        return rows, fieldnames


def write_results(rows, path):
    """Write rows to .xlsx (if the path ends in .xlsx) or to CSV otherwise."""
    if not rows:
        return
    fieldnames = list(rows[0].keys())
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(fieldnames)
        for r in rows:
            ws.append([r.get(fn, "") for fn in fieldnames])
        wb.save(path)
    else:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)


def default_out_path(ideas_path):
    ext = os.path.splitext(ideas_path)[1].lower()
    return "scored_ideas.xlsx" if ext in (".xlsx", ".xlsm") else "scored_ideas.csv"


def cmd_score(args):
    if not os.path.exists(args.model):
        sys.exit(f"Model file not found: {args.model}. Run the build step first.")
    vocab, edges, proto, meta = load_model(args.model)

    have_glove = bool(args.glove)
    if have_glove and not os.path.exists(args.glove):
        sys.exit(f"GloVe file not found: {args.glove}.")
    if args.open_only and not have_glove:
        sys.exit("--open-only needs the GloVe file too. Add --glove PATH_TO_GLOVE.")

    if args.ideas:
        rows, fieldnames = read_ideas(args.ideas)
        if not rows:
            sys.exit(f"No idea rows found in {args.ideas}.")
        cols = {c.lower(): c for c in fieldnames}
        tcol = cols.get("title")
        dcol = cols.get("description")
        if not dcol:
            sys.exit("Ideas file needs a 'Description' column, and ideally a 'Title' column.\n"
                     "Columns found: " + ", ".join(fieldnames))

        open_vectors = None
        if have_glove:
            what = "every idea" if args.open_only else "the ideas closed mode misses"
            print(f"Loading GloVe vectors to score {what} in open mode ...")
            open_vectors = load_glove(args.glove, _collect_words(rows, tcol, dcol))

        out_rows = []
        n_closed = n_open = n_unscored = 0
        for r in rows:
            idea = {"title": (r.get(tcol, "") if tcol else "") or "",
                    "description": r.get(dcol, "") or ""}
            res, mode = score_idea(idea, vocab, edges, proto, open_vectors, args.open_only)
            if mode == "closed":
                n_closed += 1
            elif mode == "open":
                n_open += 1
            else:
                n_unscored += 1

            out = dict(r)
            out["ks"] = round(res["ks"], 4) if res["scorable"] else ""
            out["prototypicality"] = round(res["prototypicality"], 4) if res["scorable"] else ""
            out["n_nodes"] = res["n_nodes"]
            out["n_edges"] = res["n_edges"]
            out["scorable"] = res["scorable"]
            out["score_mode"] = mode
            out_rows.append(out)

        out_path = args.out or default_out_path(args.ideas)
        write_results(out_rows, out_path)
        print(f"Scored {len(out_rows)} ideas. Wrote {out_path}")

        parts = []
        if not args.open_only:
            parts.append(f"{n_closed} on the topic vocabulary (closed mode)")
        if have_glove:
            parts.append(f"{n_open} from their own words (open mode)")
        parts.append(f"{n_unscored} not scorable")
        print("  " + ", ".join(parts) + ".")

        if n_unscored:
            if not have_glove:
                print(f"  {n_unscored} ideas have fewer than two words in the topic vocabulary, so")
                print("  closed mode cannot score them. They are NOT empty ideas — re-run with a GloVe")
                print("  file to score them from their own content words (open mode):")
                print(f"      python score_glove.py score --model {args.model} "
                      f"--ideas {args.ideas} --glove glove.6B.300d.txt")
                print("  That rescues every idea that has at least two real content words.")
            else:
                print(f"  The remaining {n_unscored} have fewer than two content words with a GloVe")
                print("  vector at all (for example an idea literally titled \"no\"). Nothing can score those.")
        thin = sum(1 for o in out_rows if o["scorable"] and o["n_nodes"] < 4)
        if thin:
            print(f"  {thin} ideas scored on fewer than 4 nodes, so their KPI is less stable.")
    else:
        idea = {"title": args.title or "", "description": args.description or ""}
        if not idea["description"] and not idea["title"]:
            sys.exit("Provide --title and/or --description, or --ideas a CSV or Excel file.")
        open_vectors = None
        if have_glove:
            open_vectors = load_glove(
                args.glove, _collect_words([{"t": idea["title"], "d": idea["description"]}], "t", "d"))
        res, mode = score_idea(idea, vocab, edges, proto, open_vectors, args.open_only)
        res["score_mode"] = mode
        print(json.dumps(res, indent=2))


def main():
    p = argparse.ArgumentParser(
        description="Score ideas with GloVe-based Toubia and Netzer prototypicality.")
    sub = p.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build",
                       help="Build and cache the model from topic documents and GloVe vectors.")
    b.add_argument("--glove", required=True,
                   help="Path to a GloVe text file, e.g. glove.6B.300d.txt")
    b.add_argument("--docs", help="Text file with one topic document per line.")
    b.add_argument("--docs-dir", help="Folder of .txt files, each one a topic document.")
    b.add_argument("--model", default="glove_model.json",
                   help="Where to save the model JSON. Default glove_model.json")
    b.add_argument("--min-doc-count", type=int, default=0,
                   help="Min documents a stem must appear in. Default is about 2 percent of documents "
                        "(at least 2). Lower it to keep more nodes so more ideas are scorable.")
    b.set_defaults(func=cmd_build)

    s = sub.add_parser("score",
                       help="Score one idea or a CSV of ideas using a saved model.")
    s.add_argument("--model", default="glove_model.json",
                   help="Path to the model JSON from the build step.")
    s.add_argument("--title", help="Idea title (single idea mode).")
    s.add_argument("--description", help="Idea description (single idea mode).")
    s.add_argument("--ideas", help="CSV or Excel (.xlsx) file with Title and Description columns (batch mode).")
    s.add_argument("--out", help="Output file. Use .xlsx for Excel or .csv for CSV. "
                                 "Default matches the input type.")
    s.add_argument("--glove", help="Optional path to the GloVe file. When given, any idea that closed "
                                   "mode cannot score (too few of its words are in the topic vocabulary) "
                                   "is rescued by scoring it from all of its own content words, so nearly "
                                   "every idea with two content words gets a KPI.")
    s.add_argument("--open-only", action="store_true",
                   help="With --glove, score EVERY idea from its own content words (open mode), not only "
                        "the ones closed mode misses. Use this when you want all rows scored the same way "
                        "for a fair comparison. Requires --glove.")
    s.set_defaults(func=cmd_score)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
