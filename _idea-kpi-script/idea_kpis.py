#!/usr/bin/env python3
"""
idea_kpis.py — Local, deterministic idea-ranking KPIs for the Ideation Challenge.

Run this on the command line over the aggregated research workbook produced by the
ideasearchlab Data Analytics page (``idea_analytics_aggregate.xlsx``). It reads the
``Ideas`` sheet (one row per idea, with a Title + Description) and computes the
objective, repeatable KPIs from the two spec files shipped with this study:

  • idea_ranking_kpis_llm_guide.md  (Lee & Chung 2024; Meincke et al. 2025)
        Novelty          = 1 − max cosine similarity to a reference set R
        Distinctiveness  = 1 − mean cosine similarity to the other ideas in the pool
        Combined score   = w_novelty·Novelty + w_distinct·Distinctiveness   (+ ranking)
        Unique fraction  = connected groups / N  (edge iff sim > tau)        [per pool]
  • llm_kpi_calculation_spec.md      (Bouschery et al. 2024)
        KPI 2 Productivity = count of non-redundant, multi-word ideas         [per pool]

EXCLUDED for now (by request — "complicated to compute, will add later"):
  • KPI 1  Prototypicality (KS statistic) — needs a topic web corpus + Porter stemming
           + a Jaccard semantic network + the prototypical CDF.
  • KPI 3  Brainstorming creativity — defined as the share of ideas below the KS
           creativity cutoff, so it depends on KPI 1 and is deferred together with it.

This is the offline twin of the in-app ``deterministicKpis.js`` and produces the same
numbers for the same embeddings. The pure-arithmetic KPI core is unit-tested against
the worked examples in the spec files — run ``python idea_kpis.py --selftest``.

Embeddings (text → vectors) come from one of:
  • ``--backend st``     sentence-transformers (local, recommended; e.g. all-MiniLM-L6-v2)
  • ``--backend tfidf``  a built-in deterministic TF-IDF (no extra install; APPROXIMATE,
                          lexical-overlap only — clearly labelled in the output)
  • ``--backend auto``   (default) use sentence-transformers if installed, else TF-IDF.

Quick start (Windows CMD):
    cd <repo>\\_idea-kpi-script
    pip install -r requirements.txt
    python idea_kpis.py --input idea_analytics_aggregate.xlsx
See README.md for the full guide.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import sys

import numpy as np

# ───────────────────────────────────────────────────────────────────────────────
#  1.  The KPI core — pure arithmetic over similarity values.
#      Mirrors _ideasearchlab-src/src/utils/deterministicKpis.js exactly, so the
#      offline script and the web app can never drift. Fully unit-testable: the
#      worked-example self-checks feed similarity numbers straight into these.
# ───────────────────────────────────────────────────────────────────────────────


def cosine(a, b):
    """Cosine similarity of two numeric vectors; 0.0 if either has zero length."""
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    denom = float(np.linalg.norm(a) * np.linalg.norm(b))
    return 0.0 if denom == 0.0 else float(np.dot(a, b) / denom)


def sim_matrix(vecs):
    """Full N×N cosine similarity matrix for a list of vectors (diagonal = 1).

    Returns a plain list-of-lists so every downstream consumer (and the tests) can
    treat it like the JS array. Zero-length rows give 0 similarity but a 1 on the
    diagonal, matching deterministicKpis.simMatrix.
    """
    X = np.asarray(vecs, dtype=float)
    if X.ndim != 2 or X.shape[0] == 0:
        return [[1.0]] * 0
    norms = np.linalg.norm(X, axis=1, keepdims=True)
    norms[norms == 0] = 1.0  # avoid divide-by-zero; a zero row stays the zero vector
    Xn = X / norms
    M = Xn @ Xn.T
    np.clip(M, -1.0, 1.0, out=M)  # guard tiny FP overshoot past ±1
    np.fill_diagonal(M, 1.0)
    return M.tolist()


def novelty_from_sims(sims_to_R):
    """Novelty given the list of cosine similarities to each reference-set item.

    Novelty = 1 − the single highest similarity. None if R is empty (undefined).
    """
    if not sims_to_R:
        return None
    return 1.0 - max(sims_to_R)


def novelty(idea_vec, ref_vecs):
    """Novelty of an idea = 1 − max cosine similarity to any item in R."""
    if not ref_vecs:
        return None
    return novelty_from_sims([cosine(idea_vec, r) for r in ref_vecs])


def distinctiveness_from_row(sim_row, i):
    """Pool distinctiveness of idea ``i`` from its row of the similarity matrix.

    = 1 − mean cosine similarity to the other N−1 ideas (the i-th, self, is skipped).
    None for a pool of one (the mean is undefined).
    """
    n = len(sim_row)
    if n < 2:
        return None
    total = sum(sim_row[j] for j in range(n) if j != i)
    return 1.0 - total / (n - 1)


def combined_score(nov, dist, w_nov=0.5, w_dist=0.5):
    """Combined per-idea score = w_nov·Novelty + w_dist·Distinctiveness.

    If distinctiveness is None (a pool of one) the score equals novelty (per spec);
    if novelty is None (no reference set) it falls back to distinctiveness alone.
    """
    if nov is None and dist is None:
        return None
    if dist is None:
        return nov
    if nov is None:
        return dist
    return w_nov * nov + w_dist * dist


def unique_fraction(matrix, tau=0.8):
    """Pool unique fraction = (connected groups) / N, edge iff similarity > tau.

    Two ideas join the same group when linked directly or through a chain of
    above-threshold edges (connected components via iterative DFS). Overlap is
    STRICTLY greater than tau, matching the paper. None for an empty pool.
    """
    n = len(matrix)
    if n == 0:
        return None
    seen = [False] * n
    groups = 0
    for start in range(n):
        if seen[start]:
            continue
        groups += 1
        stack = [start]
        while stack:
            node = stack.pop()
            if seen[node]:
                continue
            seen[node] = True
            row = matrix[node]
            for j in range(n):
                if not seen[j] and j != node and row[j] > tau:
                    stack.append(j)
    return groups / n


def _word_count(text):
    return len([w for w in re.split(r"\s+", str(text or "").strip()) if w])


def productivity_count(items, get_sim=None, dedup_tau=0.9, min_words=2):
    """KPI 2 — Brainstorming productivity: count of non-redundant ideas in a pool.

    Cleaning rules (Bouschery et al. 2024 §4):
      • drop empty and single-word ideas (cannot be scored),
      • within each group, collapse near-duplicates (sim > dedup_tau) into one, so the
        same solution is counted once.

    ``items``    list of dicts: {"text": str, "group": hashable}
    ``get_sim``  optional callable (i, j) -> similarity over the ORIGINAL item indices.
                 If omitted, near-duplicates are detected by exact normalised-text match.
    Returns {"count", "kept", "dropped"}.
    """
    usable = [dict(it, _i=i) for i, it in enumerate(items) if _word_count(it.get("text")) >= min_words]
    dropped = len(items) - len(usable)

    by_group = {}
    for it in usable:
        by_group.setdefault(str(it.get("group", "")), []).append(it)

    def norm(t):
        return re.sub(r"\s+", " ", str(t or "").lower()).strip()

    def near(a, b):
        if callable(get_sim):
            return get_sim(a["_i"], b["_i"]) > dedup_tau
        return norm(a["text"]) == norm(b["text"])

    count = 0
    for group in by_group.values():
        n = len(group)
        seen = [False] * n
        for s in range(n):
            if seen[s]:
                continue
            count += 1  # one fresh cluster of near-duplicates
            stack = [s]
            while stack:
                node = stack.pop()
                if seen[node]:
                    continue
                seen[node] = True
                for j in range(n):
                    if not seen[j] and j != node and near(group[node], group[j]):
                        stack.append(j)
    return {"count": count, "kept": len(usable), "dropped": dropped}


# ───────────────────────────────────────────────────────────────────────────────
#  2.  Embedding backends — text → vectors.
# ───────────────────────────────────────────────────────────────────────────────


class TfidfBackend:
    """A self-contained, deterministic TF-IDF vectoriser (no external ML model).

    APPROXIMATE: it measures lexical (word) overlap, not deeper meaning, so two ideas
    that share few words score as dissimilar even when they mean the same thing. It is
    here so the script always runs with zero heavy installs; for publication-quality,
    semantic novelty/distinctiveness use ``--backend st`` (sentence-transformers).

    Ideas and the reference set must share one vector space, so call ``fit`` once on the
    union of every text, then ``transform`` each list.
    """

    name = "tfidf (approximate, lexical-overlap only)"
    semantic = False

    _TOKEN = re.compile(r"[a-z0-9]+")

    def __init__(self):
        self.vocab = {}
        self.idf = None

    def _tokens(self, text):
        return [t for t in self._TOKEN.findall(str(text or "").lower()) if len(t) > 1]

    def fit(self, texts):
        df = {}
        for text in texts:
            for tok in set(self._tokens(text)):
                df[tok] = df.get(tok, 0) + 1
        self.vocab = {tok: i for i, tok in enumerate(sorted(df))}
        n_docs = max(1, len(texts))
        self.idf = np.zeros(len(self.vocab), dtype=float)
        for tok, i in self.vocab.items():
            # smoothed idf, identical form to scikit-learn's default
            self.idf[i] = math.log((1.0 + n_docs) / (1.0 + df[tok])) + 1.0
        return self

    def transform(self, texts):
        vecs = np.zeros((len(texts), len(self.vocab)), dtype=float)
        for r, text in enumerate(texts):
            for tok in self._tokens(text):
                j = self.vocab.get(tok)
                if j is not None:
                    vecs[r, j] += 1.0
            vecs[r] *= self.idf
            norm = np.linalg.norm(vecs[r])
            if norm > 0:
                vecs[r] /= norm
        return vecs


class STBackend:
    """sentence-transformers backend (local semantic embeddings, recommended)."""

    semantic = True

    def __init__(self, model_name="all-MiniLM-L6-v2"):
        from sentence_transformers import SentenceTransformer  # imported lazily

        self.model_name = model_name
        self.name = f"sentence-transformers/{model_name}"
        self.model = SentenceTransformer(model_name)

    def fit(self, texts):  # nothing to fit; the model is pre-trained
        return self

    def transform(self, texts):
        # normalize_embeddings=True → cosine == dot product, fully deterministic
        return np.asarray(
            self.model.encode(list(texts), normalize_embeddings=True, show_progress_bar=False),
            dtype=float,
        )


def make_backend(kind, model_name):
    """Resolve the requested backend, with the 'auto' fallback logic + clear logging."""
    if kind == "tfidf":
        return TfidfBackend()
    if kind in ("st", "auto"):
        try:
            return STBackend(model_name)
        except Exception as exc:  # not installed, or model download failed
            if kind == "st":
                print(
                    "ERROR: --backend st needs sentence-transformers.\n"
                    "       Install it with:  pip install -r requirements-semantic.txt\n"
                    f"       (import/load failed: {exc})",
                    file=sys.stderr,
                )
                sys.exit(2)
            print(
                "NOTE: sentence-transformers not available — falling back to the built-in\n"
                "      TF-IDF backend (APPROXIMATE, lexical overlap only). For semantic\n"
                "      novelty/distinctiveness run:  pip install -r requirements-semantic.txt",
                file=sys.stderr,
            )
            return TfidfBackend()
    raise ValueError(f"Unknown backend: {kind}")


# ───────────────────────────────────────────────────────────────────────────────
#  3.  Reference set R and the study's condition encoding.
# ───────────────────────────────────────────────────────────────────────────────

# Verbatim from idea_ranking_kpis_llm_guide.md §11.2 — a representative list of
# products that already exist in this market (colour-change-at-37°C fabric). Novelty
# is 1 − max similarity to these. Identical to DEFAULT_REFERENCE_SET in analyticsData.js.
DEFAULT_REFERENCE_SET = [
    "Hypercolor-style colour-change t-shirt",
    "hidden-design reveal t-shirt that shows a pattern when warmed",
    "thermochromic hoodie",
    "colour-change athletic top",
    "thermochromic socks",
    "colour-changing swim shorts",
    "mood ring",
    "mood necklace",
    "thermochromic bracelet or beads",
    "thermochromic phone case",
    "thermochromic nail polish",
    "colour-change lipstick",
    "photochromic eyeglass lenses",
    "forehead fever thermometer strip",
    "thermochromic fever-indicator baby sticker",
    "colour-changing baby feeding spoon",
    "thermochromic baby bath thermometer or toy",
    "liquid-crystal room or aquarium strip",
    "colour-changing coffee mug",
    "thermochromic kettle band",
    "colour-change bath or floor mat",
    "thermochromic shower-head indicator",
]

VALID_CONDITIONS = {"None", "Solo", "Group", "Both"}


def condition_from_flags(solo, group):
    """AI-Solo × AI-Group dummies → the placement encoding (matches conditionFromFlags)."""
    if solo and group:
        return "Both"   # Full AI
    if solo and not group:
        return "Solo"   # Individual + AI
    if not solo and group:
        return "Group"  # Group + AI
    return "None"       # Human-Only Hybrid


def idea_text(title, description):
    """Combined scoring text for an idea (matches analyticsData.ideaText)."""
    title = (title or "").strip()
    description = (description or "").strip()
    if title and description:
        return f"{title}: {description}"
    return title or description


# ───────────────────────────────────────────────────────────────────────────────
#  4.  Reading the workbook.
# ───────────────────────────────────────────────────────────────────────────────


def _truthy_flag(v):
    return str(v or "").strip() in ("1", "1.0", "yes", "true", "True")


def load_ideas(path, sheet_name="Ideas"):
    """Read the Ideas sheet into a list of idea dicts, mapping columns by header text.

    Drops rows flagged ``Exclude (Yes/No) = Yes`` (the pre-registered screen) and rows
    whose combined Title+Description is blank. Returns (ideas, meta).
    """
    try:
        import openpyxl
    except ImportError:
        print(
            "ERROR: openpyxl is required to read .xlsx files.\n"
            "       Install it with:  pip install -r requirements.txt",
            file=sys.stderr,
        )
        sys.exit(2)

    if not os.path.exists(path):
        print(f"ERROR: input file not found: {path}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        print(
            f"ERROR: sheet '{sheet_name}' not found. Sheets in this file: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(2)
    ws = wb[sheet_name]

    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        print(f"ERROR: sheet '{sheet_name}' is empty.", file=sys.stderr)
        sys.exit(2)

    col = {}
    for i, name in enumerate(header):
        if name is not None:
            col[str(name).strip()] = i

    def get(row, name, default=""):
        i = col.get(name)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return default if v is None else v

    if "Title" not in col and "Description" not in col:
        print(
            f"ERROR: the '{sheet_name}' sheet has no Title/Description columns. "
            f"Found headers: {list(col)[:12]}...",
            file=sys.stderr,
        )
        sys.exit(2)

    ideas = []
    excluded = 0
    blank = 0
    for r in rows:
        if r is None or not any(c is not None for c in r):
            continue
        if str(get(r, "Exclude (Yes/No)")).strip().lower() == "yes":
            excluded += 1
            continue
        title = str(get(r, "Title")).strip()
        description = str(get(r, "Description")).strip()
        text = idea_text(title, description)
        if not text:
            blank += 1
            continue

        condition = str(get(r, "Condition")).strip()
        if condition not in VALID_CONDITIONS:
            condition = condition_from_flags(
                _truthy_flag(get(r, "AI Solo (0/1)")), _truthy_flag(get(r, "AI Group (0/1)"))
            )

        group_uid = str(get(r, "Group UID")).strip() or str(get(r, "Group ID")).strip()
        ideas.append(
            {
                "idea_id": str(get(r, "Idea ID")).strip(),
                "session": str(get(r, "Session Code")).strip(),
                "condition": condition,
                "stage": str(get(r, "Stage")).strip(),
                "group_uid": group_uid,
                "author_label": str(get(r, "Author Label")).strip(),
                "title": title,
                "description": description,
                "text": text,
            }
        )
    wb.close()
    meta = {"rows_excluded_flag": excluded, "rows_blank": blank, "ideas_loaded": len(ideas)}
    return ideas, meta


def load_reference_set(path):
    """Reference set R: one product per line from ``path``; blank/`#` lines ignored.

    Falls back to the built-in 22-item DEFAULT_REFERENCE_SET when no file is given or
    the file is missing.
    """
    if not path:
        return list(DEFAULT_REFERENCE_SET), "built-in default (spec §11.2, 22 items)"
    if not os.path.exists(path):
        print(
            f"NOTE: reference file '{path}' not found — using the built-in default R.",
            file=sys.stderr,
        )
        return list(DEFAULT_REFERENCE_SET), "built-in default (spec §11.2, 22 items)"
    with open(path, "r", encoding="utf-8") as fh:
        items = [ln.strip() for ln in fh if ln.strip() and not ln.lstrip().startswith("#")]
    if not items:
        print(f"NOTE: reference file '{path}' is empty — using the built-in default R.", file=sys.stderr)
        return list(DEFAULT_REFERENCE_SET), "built-in default (spec §11.2, 22 items)"
    return items, f"file: {os.path.basename(path)} ({len(items)} items)"


# ───────────────────────────────────────────────────────────────────────────────
#  5.  Orchestration — compute every in-scope KPI over the loaded ideas.
# ───────────────────────────────────────────────────────────────────────────────


def compute_kpis(ideas, refs, backend, *, pool_by="", tau=0.8, w_nov=0.5, w_dist=0.5,
                 taus=(0.75, 0.8, 0.85), dedup_tau=0.9, min_words=2):
    """Compute per-idea and per-pool KPIs. Returns a result dict (JSON-serialisable).

    Pools are defined by the ``pool_by`` idea field ("" → the whole file is one pool).
    Distinctiveness, the ranking and the per-pool measures all use that grouping.
    """
    texts = [it["text"] for it in ideas]

    # Embed ideas + R in one shared vector space (so novelty similarities are valid).
    backend.fit(texts + refs)
    idea_vecs = backend.transform(texts)
    ref_vecs = backend.transform(refs) if refs else np.zeros((0, idea_vecs.shape[1] if idea_vecs.size else 1))

    # Novelty (vs R) is independent of the pool — vectorise it for all ideas at once.
    if len(refs):
        ideas_to_R = idea_vecs @ ref_vecs.T  # both already L2-normalised → cosine
        max_sim_to_R = ideas_to_R.max(axis=1)
        novelties = [1.0 - float(m) for m in max_sim_to_R]
    else:
        novelties = [None] * len(ideas)

    # Group idea indices into pools.
    pools = {}
    for i, it in enumerate(ideas):
        key = it.get(pool_by, "") if pool_by else "(all ideas)"
        pools.setdefault(str(key) or "(blank)", []).append(i)

    # Per-idea distinctiveness + combined score, computed within each pool.
    dist = [None] * len(ideas)
    score = [None] * len(ideas)
    for idxs in pools.values():
        M = sim_matrix([idea_vecs[i] for i in idxs])
        for local_i, gi in enumerate(idxs):
            d = distinctiveness_from_row(M[local_i], local_i)
            dist[gi] = d
            score[gi] = combined_score(novelties[gi], d, w_nov, w_dist)

    for i, it in enumerate(ideas):
        it["novelty"] = novelties[i]
        it["distinctiveness"] = dist[i]
        it["score"] = score[i]
        it["pool"] = (str(it.get(pool_by, "")) or "(blank)") if pool_by else "(all ideas)"

    # Ranking within each pool: by score desc, tie-break novelty → distinctiveness →
    # original order (spec §7). Non-scored ideas (no score) rank last.
    for idxs in pools.values():
        def sort_key(i):
            s = ideas[i]["score"]
            nv = ideas[i]["novelty"]
            dv = ideas[i]["distinctiveness"]
            return (
                s if s is not None else -math.inf,
                nv if nv is not None else -math.inf,
                dv if dv is not None else -math.inf,
            )
        for rank, gi in enumerate(sorted(idxs, key=sort_key, reverse=True), start=1):
            ideas[gi]["rank_in_pool"] = rank

    # Per-pool measures: unique fraction (3 taus) + KPI 2 productivity.
    pool_results = []
    for pool_key, idxs in pools.items():
        vecs = [idea_vecs[i] for i in idxs]
        M = sim_matrix(vecs)
        local_index = {gi: k for k, gi in enumerate(idxs)}
        items = [{"text": ideas[i]["text"], "group": ideas[i]["group_uid"]} for i in idxs]

        def get_sim(a, b, _M=M):  # a,b are positions within this pool's item list
            return _M[a][b]

        prod = productivity_count(items, get_sim, dedup_tau=dedup_tau, min_words=min_words)
        pool_results.append(
            {
                "pool": pool_key,
                "n_ideas": len(idxs),
                "productivity_count": prod["count"],
                "productivity_dropped_short": prod["dropped"],
                "unique_fraction": {f"tau_{t}": unique_fraction(M, t) for t in taus},
            }
        )
    pool_results.sort(key=lambda p: (-p["n_ideas"], p["pool"]))

    return {
        "settings": {
            "embedding_backend": backend.name,
            "semantic_embeddings": bool(getattr(backend, "semantic", False)),
            "reference_set_size": len(refs),
            "pool_by": pool_by or "(whole file as one pool)",
            "tau": tau,
            "weights": {"novelty": w_nov, "distinctiveness": w_dist},
            "unique_fraction_taus": list(taus),
            "productivity": {"dedup_tau": dedup_tau, "min_words": min_words},
            "excluded_kpis": [
                "KPI 1 Prototypicality (KS statistic) — deferred",
                "KPI 3 Brainstorming creativity — depends on KPI 1, deferred with it",
            ],
        },
        "ideas": ideas,
        "pools": pool_results,
    }


# ───────────────────────────────────────────────────────────────────────────────
#  6.  Output — console summary + CSV + JSON.
# ───────────────────────────────────────────────────────────────────────────────


def _fmt(x, nd=4):
    return "" if x is None else f"{x:.{nd}f}"


def write_outputs(result, outdir, basename="idea_kpis"):
    os.makedirs(outdir, exist_ok=True)
    ideas = result["ideas"]

    per_idea = os.path.join(outdir, f"{basename}_per_idea.csv")
    fields = [
        "rank_in_pool", "pool", "idea_id", "session", "condition", "stage", "group_uid",
        "author_label", "title", "description", "novelty", "distinctiveness", "score",
    ]
    with open(per_idea, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(fields)
        for it in sorted(ideas, key=lambda r: (str(r["pool"]), r.get("rank_in_pool", 1e9))):
            w.writerow(
                [
                    it.get("rank_in_pool", ""), it["pool"], it["idea_id"], it["session"],
                    it["condition"], it["stage"], it["group_uid"], it["author_label"],
                    it["title"], it["description"],
                    _fmt(it["novelty"]), _fmt(it["distinctiveness"]), _fmt(it["score"]),
                ]
            )

    per_pool = os.path.join(outdir, f"{basename}_pools.csv")
    taus = result["settings"]["unique_fraction_taus"]
    with open(per_pool, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(
            ["pool", "n_ideas", "productivity_count", "productivity_dropped_short"]
            + [f"unique_fraction_tau_{t}" for t in taus]
        )
        for p in result["pools"]:
            w.writerow(
                [p["pool"], p["n_ideas"], p["productivity_count"], p["productivity_dropped_short"]]
                + [_fmt(p["unique_fraction"][f"tau_{t}"], 4) for t in taus]
            )

    js = os.path.join(outdir, f"{basename}.json")
    with open(js, "w", encoding="utf-8") as fh:
        json.dump(result, fh, indent=2, ensure_ascii=False)

    return per_idea, per_pool, js


def print_summary(result, top=15):
    s = result["settings"]
    ideas = result["ideas"]
    print("\n" + "=" * 78)
    print("  IDEA-RANKING KPIs  (deterministic; KPI 1 KS prototypicality excluded)")
    print("=" * 78)
    print(f"  Embedding backend : {s['embedding_backend']}")
    if not s["semantic_embeddings"]:
        print("                      ^ APPROXIMATE (lexical overlap). Install")
        print("                        sentence-transformers for semantic embeddings.")
    print(f"  Ideas scored      : {len(ideas)}")
    print(f"  Reference set R   : {s['reference_set_size']} products")
    print(f"  Pool grouping     : {s['pool_by']}")
    print(f"  Weights / tau     : novelty {s['weights']['novelty']}, "
          f"distinct {s['weights']['distinctiveness']} / tau {s['tau']}")

    print("\n  POOL-LEVEL KPIs")
    taus = s["unique_fraction_taus"]
    header = f"  {'pool':<22}{'n':>5}{'KPI2 prod.':>12}" + "".join(f"{'UF@'+str(t):>10}" for t in taus)
    print(header)
    print("  " + "-" * (len(header) - 2))
    for p in result["pools"]:
        uf = "".join(f"{_fmt(p['unique_fraction'][f'tau_{t}'], 2):>10}" for t in taus)
        print(f"  {p['pool'][:22]:<22}{p['n_ideas']:>5}{p['productivity_count']:>12}{uf}")

    # Headline ranking: best ideas across all pools by combined score.
    scored = [it for it in ideas if it["score"] is not None]
    scored.sort(key=lambda r: r["score"], reverse=True)
    print(f"\n  TOP {min(top, len(scored))} IDEAS BY COMBINED SCORE (Novelty + Distinctiveness)")
    print(f"  {'#':>3}  {'score':>6} {'novl':>6} {'dist':>6}  {'pool':<10} title")
    print("  " + "-" * 74)
    for i, it in enumerate(scored[:top], start=1):
        print(
            f"  {i:>3}  {_fmt(it['score'], 3):>6} {_fmt(it['novelty'], 3):>6} "
            f"{_fmt(it['distinctiveness'], 3):>6}  {str(it['pool'])[:10]:<10} {it['title'][:40]}"
        )
    print("=" * 78 + "\n")


# ───────────────────────────────────────────────────────────────────────────────
#  7.  Self-tests — reproduce the documented worked examples (no embeddings needed).
# ───────────────────────────────────────────────────────────────────────────────


def run_selftest():
    """Assert the KPI core against the worked examples in the spec files."""
    failures = []

    def check(name, got, want, tol=5e-3):
        ok = (got is None and want is None) or (
            got is not None and want is not None and abs(got - want) <= tol
        )
        print(f"  [{'PASS' if ok else 'FAIL'}] {name}: got {got}, want {want}")
        if not ok:
            failures.append(name)

    print("\n--- cosine sanity ---")
    check("cosine identical", cosine([1, 0, 1], [1, 0, 1]), 1.0)
    check("cosine orthogonal", cosine([1, 0], [0, 1]), 0.0)
    check("cosine zero-vector", cosine([0, 0], [1, 1]), 0.0)
    check("cosine scaled", cosine([2, 0], [5, 0]), 1.0)

    # idea_ranking_kpis_llm_guide.md §11.3 — Novelty (1 − max similarity to R).
    print("\n--- Novelty (guide §11.3) ---")
    tee_sims = [0.71, 0.79, 0.66, 0.52, 0.38, 0.33, 0.44, 0.41, 0.39, 0.31, 0.27,
                0.25, 0.36, 0.38, 0.34, 0.22, 0.24, 0.28, 0.29, 0.26, 0.30, 0.25]
    check("novelty Reveal Tee", novelty_from_sims(tee_sims), 0.21)
    check("novelty Sports Bra", novelty_from_sims([0.62, 0.55, 0.50]), 0.38)
    check("novelty Baby Onesie", novelty_from_sims([0.74, 0.61, 0.48]), 0.26)
    check("novelty Leggings", novelty_from_sims([0.58, 0.46, 0.34]), 0.42)
    check("novelty Yoga Mat", novelty_from_sims([0.47, 0.38, 0.33]), 0.53)
    check("novelty empty R", novelty_from_sims([]), None)

    # §11.4 — Pool distinctiveness from the 5×5 idea similarity matrix.
    print("\n--- Distinctiveness (guide §11.4) ---")
    S = [
        [1.00, 0.82, 0.41, 0.53, 0.31],
        [0.82, 1.00, 0.39, 0.58, 0.34],
        [0.41, 0.39, 1.00, 0.36, 0.22],
        [0.53, 0.58, 0.36, 1.00, 0.44],
        [0.31, 0.34, 0.22, 0.44, 1.00],
    ]
    # Exact pre-rounded values from the spec's own arithmetic (§11.4). The guide's
    # table displays these to 2 dp (e.g. 0.655 → "0.66"), so we assert the exact
    # values with a tight tolerance rather than the 2-dp display.
    want_dist = [0.4825, 0.4675, 0.655, 0.5225, 0.6725]
    for i, w in enumerate(want_dist):
        check(f"distinctiveness idea {i + 1}", distinctiveness_from_row(S[i], i), w, tol=1e-9)
    check("distinctiveness pool-of-one", distinctiveness_from_row([1.0], 0), None)

    # §11.5 — Unique fraction at three thresholds (only pair (1,2)=0.82 overlaps).
    print("\n--- Unique fraction (guide §11.5) ---")
    check("unique_fraction tau=0.80", unique_fraction(S, 0.80), 0.80)
    check("unique_fraction tau=0.75", unique_fraction(S, 0.75), 0.80)
    check("unique_fraction tau=0.85", unique_fraction(S, 0.85), 1.00)

    # §6.3 — the paper's two worked unique-fraction checks.
    print("\n--- Unique fraction (guide §6.3 worked checks) ---")
    distinct3 = [[1.0, 0.1, 0.2], [0.1, 1.0, 0.15], [0.2, 0.15, 1.0]]  # all separate
    check("UF {furniture,rack,mount}", unique_fraction(distinct3, 0.8), 1.00)
    same3 = [[1.0, 0.9, 0.88], [0.9, 1.0, 0.91], [0.88, 0.91, 1.0]]    # all the same
    check("UF {3 sprinklers}", unique_fraction(same3, 0.8), 1 / 3)

    # §11.6 / §7 — combined score and ranking.
    print("\n--- Combined score + ranking (guide §11.6–§11.7) ---")
    nov = [0.21, 0.38, 0.26, 0.42, 0.53]
    dst = [0.48, 0.47, 0.66, 0.52, 0.67]
    titles = ["Reveal Tee", "Sports Bra", "Baby Onesie", "Leggings", "Yoga Mat"]
    scores = [combined_score(nov[i], dst[i]) for i in range(5)]
    for i, w in enumerate([0.345, 0.425, 0.46, 0.47, 0.60]):
        check(f"score {titles[i]}", scores[i], w)
    order = [titles[i] for i in sorted(range(5), key=lambda i: scores[i], reverse=True)]
    want_order = ["Yoga Mat", "Leggings", "Baby Onesie", "Sports Bra", "Reveal Tee"]
    print(f"  [{'PASS' if order == want_order else 'FAIL'}] ranking order: {order}")
    if order != want_order:
        failures.append("ranking order")

    # KPI 2 productivity (Bouschery §4): drop short ideas; merge near-duplicates per group.
    print("\n--- KPI 2 Productivity (spec §4) ---")
    prod_items = [
        {"text": "Garden furniture set", "group": "g1"},
        {"text": "Storage rack system", "group": "g1"},
        {"text": "Picture mount frame", "group": "g1"},
        {"text": "medical", "group": "g1"},              # single word → dropped
        {"text": "  ", "group": "g1"},                   # blank → dropped
    ]
    p1 = productivity_count(prod_items)  # no get_sim → exact-text dedup; all 3 distinct
    check("productivity distinct-3 count", float(p1["count"]), 3.0)
    check("productivity dropped-short", float(p1["dropped"]), 2.0)

    sprinklers = [
        {"text": "Lawn sprinkler", "group": "g1"},
        {"text": "Water sprinkler", "group": "g1"},
        {"text": "Garden sprinkler", "group": "g1"},
    ]
    sim = [[1.0, 0.95, 0.93], [0.95, 1.0, 0.92], [0.93, 0.92, 1.0]]  # all near-duplicates
    p2 = productivity_count(sprinklers, lambda a, b: sim[a][b], dedup_tau=0.9)
    check("productivity near-dup merge → 1", float(p2["count"]), 1.0)

    # near-duplicates only merge WITHIN a group, never across groups.
    cross = [
        {"text": "Lawn sprinkler", "group": "g1"},
        {"text": "Lawn sprinkler", "group": "g2"},
    ]
    p3 = productivity_count(cross, lambda a, b: 0.99, dedup_tau=0.9)
    check("productivity cross-group not merged → 2", float(p3["count"]), 2.0)

    print("\n" + ("ALL SELF-TESTS PASSED ✓" if not failures else f"FAILURES: {failures}"))
    return 0 if not failures else 1


# ───────────────────────────────────────────────────────────────────────────────
#  8.  CLI.
# ───────────────────────────────────────────────────────────────────────────────


def main(argv=None):
    here = os.path.dirname(os.path.abspath(__file__))
    parser = argparse.ArgumentParser(
        description="Compute deterministic idea-ranking KPIs (KPI 1 KS excluded).",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--input", "-i", help="Path to idea_analytics_aggregate.xlsx")
    parser.add_argument("--sheet", default="Ideas", help="Worksheet with the ideas")
    parser.add_argument("--reference", "-r", default=os.path.join(here, "reference_set.txt"),
                        help="Reference set R file (one product per line)")
    parser.add_argument("--backend", choices=["auto", "st", "tfidf"], default="auto",
                        help="Embedding backend")
    parser.add_argument("--model", default="all-MiniLM-L6-v2",
                        help="sentence-transformers model name (for --backend st/auto)")
    parser.add_argument("--pool-by", default="",
                        help="Idea field defining a pool for distinctiveness/UF/productivity "
                             "(e.g. 'session' or 'condition'); blank = whole file as one pool")
    parser.add_argument("--tau", type=float, default=0.8, help="Unique-fraction overlap threshold")
    parser.add_argument("--w-novelty", type=float, default=0.5, help="Score weight on novelty")
    parser.add_argument("--w-distinct", type=float, default=0.5, help="Score weight on distinctiveness")
    parser.add_argument("--dedup-tau", type=float, default=0.9, help="KPI 2 near-duplicate threshold")
    parser.add_argument("--min-words", type=int, default=2, help="KPI 2 minimum words to keep an idea")
    parser.add_argument("--outdir", "-o", default=os.path.join(here, "output"),
                        help="Directory for the CSV/JSON outputs")
    parser.add_argument("--top", type=int, default=15, help="How many top ideas to print")
    parser.add_argument("--selftest", action="store_true",
                        help="Run the worked-example self-checks and exit")
    args = parser.parse_args(argv)

    if args.selftest:
        return run_selftest()

    if not args.input:
        parser.error("--input is required (path to the .xlsx). Use --selftest to verify the maths.")

    ideas, meta = load_ideas(args.input, args.sheet)
    if not ideas:
        print("ERROR: no usable ideas found after cleaning.", file=sys.stderr)
        return 1
    print(
        f"Loaded {meta['ideas_loaded']} ideas "
        f"(skipped {meta['rows_excluded_flag']} Exclude-flagged, {meta['rows_blank']} blank)."
    )

    refs, ref_desc = load_reference_set(args.reference)
    print(f"Reference set R: {ref_desc}")

    pool_by = args.pool_by.strip()
    if pool_by and pool_by not in ideas[0]:
        valid = [k for k in ("session", "condition", "stage", "group_uid") if k in ideas[0]]
        parser.error(f"--pool-by '{pool_by}' is not an idea field. Try one of: {valid}, or leave blank.")

    backend = make_backend(args.backend, args.model)
    print(f"Embedding with: {backend.name} …")

    result = compute_kpis(
        ideas, refs, backend,
        pool_by=pool_by, tau=args.tau, w_nov=args.w_novelty, w_dist=args.w_distinct,
        dedup_tau=args.dedup_tau, min_words=args.min_words,
    )

    print_summary(result, top=args.top)
    per_idea, per_pool, js = write_outputs(result, args.outdir)
    print("Wrote:")
    print(f"  • {per_idea}   (one row per idea: novelty, distinctiveness, score, rank)")
    print(f"  • {per_pool}   (one row per pool: KPI 2 productivity + unique fraction)")
    print(f"  • {js}   (full structured results)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
