"""
Script 2: Fill Editor & Area from INFORMS article pages.

Supports both .csv and .xlsx/.xls input files.
Looks for: mnsc_articles_editors.csv/xlsx → mnsc_articles.csv/xlsx

Install: pip install cloudscraper openpyxl
"""

import cloudscraper, csv, time, re, sys, os, glob, random

BASE_NAME = "mnsc_articles"
OUTPUT_NAME = "mnsc_articles_editors"
BASE_URL = "https://pubsonline.informs.org/doi/abs/"

session = cloudscraper.create_scraper()


def extract_editor(html):
    # Step 1: Find "accepted by" and grab a chunk of text after it
    match = re.search(r"accepted by\s+(.{5,200})", html, re.IGNORECASE | re.DOTALL)
    if not match:
        return "", ""

    raw = match.group(1)

    # Step 2: Strip HTML tags
    raw = re.sub(r"<[^>]+>", " ", raw)
    raw = re.sub(r"\s+", " ", raw).strip()

    # Step 3: Remove leading "by " if duplicated
    raw = re.sub(r"^by\s+", "", raw, flags=re.IGNORECASE)

    # Step 4: Take everything up to the first period that ends the sentence
    # (not a period in initials like "J." or "B.")
    # Strategy: split on comma first — editor is before comma, area is after
    comma_pos = raw.find(",")
    if comma_pos == -1:
        return "", ""

    editor = raw[:comma_pos].strip()
    
    # Area: everything after comma, up to the first period followed by space/end
    rest = raw[comma_pos + 1:].strip()
    # Take until first sentence-ending period (period followed by space, end, or quote)
    area_match = re.match(r"([^<]+?)\.(?:\s|$|\"|\)|<)", rest + " ")
    area = area_match.group(1).strip() if area_match else rest.split(".")[0].strip()

    return editor, area


def scrape_article(doi_url):
    doi_id = doi_url.replace("https://doi.org/", "")
    url = BASE_URL + doi_id

    for attempt in range(10):
        try:
            resp = session.get(url, timeout=30)
            if resp.status_code == 200:
                # Check it's a real page, not a Cloudflare challenge
                if "accepted by" not in resp.text.lower() and "abstractSection" not in resp.text:
                    if "challenge" in resp.text.lower() or len(resp.text) < 5000:
                        # Cloudflare challenge disguised as 200
                        if attempt < 9:
                            wait = random.randint(5, 20)
                            print(f"(blocked, retry {attempt+1}/10 in {wait}s)", end=" ", flush=True)
                            time.sleep(wait)
                            continue
                editor, area = extract_editor(resp.text)
                return editor, area, ""
            elif resp.status_code == 403 and attempt < 9:
                wait = random.randint(3, 15)
                print(f"(403, retry {attempt+1}/10 in {wait}s)", end=" ", flush=True)
                time.sleep(wait)
                continue
            else:
                return "", "", f"HTTP {resp.status_code}"
        except Exception as e:
            if attempt < 9:
                time.sleep(random.randint(3, 10))
                continue
            return "", "", str(e)
    return "", "", "max retries"


# ── File I/O: supports CSV and Excel ─────────────────────────

def find_input_file():
    """Find the best input file: prefer editors output, then original."""
    for name in [OUTPUT_NAME, BASE_NAME]:
        for ext in [".xlsx", ".csv", ".xls"]:
            path = name + ext
            if os.path.exists(path):
                return path
    return None


def read_file(path):
    """Read CSV or Excel file, return (rows as list of dicts, fieldnames)."""
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            return rows, list(reader.fieldnames)

    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            print("Error: openpyxl is required for Excel files.")
            print("Run: pip install openpyxl")
            sys.exit(1)

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        wb.close()

        if not data:
            return [], []

        # First row = headers — only keep columns with real header names
        raw_headers = list(data[0])
        headers = []
        for h in raw_headers:
            if h and str(h).strip():
                headers.append(str(h).strip())
            else:
                break  # stop at first empty column
        num_cols = len(headers)

        rows = []
        for row_data in data[1:]:
            row_dict = {}
            for i in range(num_cols):
                val = row_data[i] if i < len(row_data) else None
                row_dict[headers[i]] = str(val) if val is not None else ""
            rows.append(row_dict)
        return rows, headers

    else:
        print(f"Error: unsupported file format '{ext}'")
        sys.exit(1)


def write_file(path, rows, fieldnames):
    """Write to xlsx (modern Excel format)."""
    import openpyxl

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        # Headers
        for c, h in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for r, row in enumerate(rows, 2):
            for c, h in enumerate(fieldnames, 1):
                ws.cell(row=r, column=c, value=row.get(h, ""))

        wb.save(path)
        wb.close()
        return path
    except PermissionError:
        for n in range(1, 100):
            alt = path.replace(".xlsx", f"_{n}.xlsx")
            try:
                wb.save(alt)
                wb.close()
                print(f"\n⚠ '{path}' is open. Saved to '{alt}' instead.")
                return alt
            except PermissionError:
                continue
    return None


# ── Main ──────────────────────────────────────────────────────

def main():
    source = find_input_file()

    if not source:
        print("Error: No input file found.")
        print(f"Place '{BASE_NAME}.csv' or '{BASE_NAME}.xlsx' in this folder.")
        print(f"Files in current directory:")
        for f in sorted(os.listdir(".")):
            print(f"  {f}")
        sys.exit(1)

    print(f"Reading: {source}")
    rows, fieldnames = read_file(source)

    if "Accepting Editor" not in fieldnames:
        fieldnames.append("Accepting Editor")
    if "Area" not in fieldnames:
        fieldnames.append("Area")

    # Ensure all rows have the new columns
    for row in rows:
        if "Accepting Editor" not in row:
            row["Accepting Editor"] = ""
        if "Area" not in row:
            row["Area"] = ""

    total = len(rows)
    already_filled = sum(1 for r in rows if r.get("Accepting Editor", "").strip())
    print(f"Loaded {total} rows")
    print(f"  Already have editor info: {already_filled}")
    print(f"  Row 1 = first data row (not header)")
    print()

    start_input = input(f"Start from row (1-{total}, default 1): ").strip()
    start_row = int(start_input) if start_input else 1

    end_input = input(f"End at row (1-{total}, default {total}): ").strip()
    end_row = int(end_input) if end_input else total

    start_row = max(1, min(start_row, total))
    end_row = max(start_row, min(end_row, total))

    print(f"\nProcessing rows {start_row} to {end_row} ({end_row - start_row + 1} rows)")
    print("Ctrl+C once = skip current row | Ctrl+C twice quickly = stop & save")
    print("=" * 60)

    filled = 0
    skipped = 0
    failed = 0
    errors = []
    stopped_at = None

    try:
        for i in range(start_row - 1, end_row):
            row = rows[i]
            row_num = i + 1
            doi = row.get("DOI", "").strip()
            status = row.get("Status", "").strip()
            existing_editor = row.get("Accepting Editor", "").strip()

            if existing_editor or not doi or status == "Other":
                skipped += 1
                continue

            title_short = row.get("Title", "")[:50]
            print(f"[Row {row_num}/{end_row}] {title_short}...", end=" ", flush=True)

            try:
                editor, area, err = scrape_article(doi)
            except KeyboardInterrupt:
                print(" ⏭ SKIPPED (Ctrl+C)")
                skipped += 1
                time.sleep(1)
                continue

            if err:
                print(f"✗ {err}")
                errors.append(f"Row {row_num}: {doi} — {err}")
                failed += 1
            elif editor:
                row["Accepting Editor"] = editor
                row["Area"] = area
                print(f"→ {editor}, {area}")
                filled += 1
            else:
                print("→ (no editor found)")
                failed += 1

            time.sleep(2)

    except KeyboardInterrupt:
        stopped_at = row_num
        print(f"\n\n⛔ STOPPED at row {row_num}. Saving progress...")

    # Output as xlsx
    out_path = OUTPUT_NAME + ".xlsx"
    saved = write_file(out_path, rows, fieldnames)

    print()
    print("=" * 60)
    print(f"Results for rows {start_row}-{end_row}:")
    print(f"  Filled:  {filled}")
    print(f"  Skipped: {skipped} (already filled, Other, or no DOI)")
    print(f"  Failed:  {failed}")
    print(f"\nOutput saved to: {saved}")

    if errors:
        print(f"\nErrors ({len(errors)}):")
        for e in errors[:10]:
            print(f"  {e}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more")

    if stopped_at:
        print(f"\n⛔ Stopped early. To resume: run again with start row = {stopped_at}")
    else:
        print(f"\nTo continue: run again with start row = {end_row + 1}")


if __name__ == "__main__":
    main()