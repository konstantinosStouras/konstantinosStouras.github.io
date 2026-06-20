"""
Script 2: Fill Editor & Area from INFORMS article pages.
Falls back to Crossref API abstract when INFORMS blocks (403).

Supports both .csv and .xlsx/.xls input files.
Looks for: mnsc_articles_editors.csv/xlsx → mnsc_articles.csv/xlsx

Install: pip install cloudscraper openpyxl requests
"""

import cloudscraper, csv, time, re, sys, os, random, json
try:
    import requests as req_lib
except ImportError:
    req_lib = None

BASE_NAME = "mnsc_articles"
OUTPUT_NAME = "mnsc_articles_editors"
BASE_URL = "https://pubsonline.informs.org/doi/abs/"
CROSSREF_API = "https://api.crossref.org/works/"
EMAIL = "kstouras@gmail.com"

session = cloudscraper.create_scraper()

# ── Titles to skip (no editor info expected) ──────────────────
SKIP_PATTERNS = [
    "management insights", "editorial", "from the editor",
    "erratum", "corrigendum", "correction to", "retraction",
    "introduction to the special issue", "introduction to the special section",
    "letter to the editor", "letter from the editor",
    "editor's comments", "note from the editor",
    "special issue on", "special section of",
    "reviewer", "referees", "index to volume",
    "service award", "best paper award", "best ae award",
    "acknowledgment", "call for papers", "announcement", "errata",
    "in memoriam", "pioneering contributions",
    "reengineering management science",
    "reinforcing research transparency",
    "management science: the legacy",
    "advances in blockchain and crypto economics",
    "msom society best", "information systems best ae",
    "management science special section",
]

def should_skip_title(title):
    """Check if a paper title indicates non-research content (no editor)."""
    lower = title.lower()
    for pat in SKIP_PATTERNS:
        if pat in lower:
            return True
    return False


def extract_editor(html):
    """Extract editor and area from HTML page text."""
    match = re.search(r"accepted by\s+(.{5,300})", html, re.IGNORECASE | re.DOTALL)
    if not match:
        # Also try "served as the editor" pattern
        match2 = re.search(r"(\w[\w\s.]+?)\s+served as (?:the )?editor", html, re.IGNORECASE)
        if match2:
            return match2.group(1).strip(), ""
        return "", ""

    raw = match.group(1)
    raw = re.sub(r"<[^>]+>", " ", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    raw = re.sub(r"^by\s+", "", raw, flags=re.IGNORECASE)

    return split_editor_area(raw)


def extract_editor_from_abstract(abstract_text):
    """Extract editor and area from Crossref abstract text."""
    if not abstract_text:
        return "", ""
    match = re.search(r"(?:This paper|This work) was accepted by\s+([^.]+(?:\.[^.]{0,5})?[^.]*)\.", abstract_text, re.IGNORECASE)
    if not match:
        match = re.search(r"accepted by\s+([^.]+(?:\.[^.]{0,5})?[^.]*)\.", abstract_text, re.IGNORECASE)
    if not match:
        # "served as the editor" pattern
        match2 = re.search(r"(\w[\w\s.]+?)\s+served as (?:the )?editor", abstract_text, re.IGNORECASE)
        if match2:
            return match2.group(1).strip(), ""
        return "", ""

    raw = match.group(1).strip()
    return split_editor_area(raw)


def split_editor_area(raw):
    """Split a raw 'editor, area' or 'editor for the Special Issue...' string."""
    # Pattern 1: "Name for the Special Issue/Section on/of ..."
    for_match = re.match(r"(.+?)\s+for the\s+(Special (?:Issue|Section).+)", raw, re.IGNORECASE)
    if for_match:
        editor = for_match.group(1).strip().rstrip(",")
        area = for_match.group(2).strip().rstrip(",").rstrip(".")
        return editor, clean_area(area)

    # Pattern 2: "Name served as the editor for this article"
    served_match = re.match(r"(.+?)\s+served as", raw, re.IGNORECASE)
    if served_match:
        return served_match.group(1).strip(), ""

    # Pattern 3: Standard "Name, area"
    comma_pos = raw.find(",")
    if comma_pos == -1:
        return raw.strip().rstrip("."), ""

    editor = raw[:comma_pos].strip()
    rest = raw[comma_pos + 1:].strip()
    area_match = re.match(r"([^<]+?)\.(?:\s|$|\"|\)|<)", rest + " ")
    area = area_match.group(1).strip() if area_match else rest.split(".")[0].strip()

    return editor, clean_area(area)


def clean_area(area):
    """Remove trailing junk from area strings."""
    if not area:
        return area
    # Truncate at known junk patterns
    for junk in ['Funding:', 'Supplemental Material:', 'Conflict of Interest',
                 'The online appendix', 'Data and the online', 'History:',
                 'https://doi.org', 'Disclaimer']:
        idx = area.find(junk)
        if idx > 0:
            area = area[:idx]
    # Remove trailing period/comma/whitespace
    area = area.strip().rstrip('.,;: ')
    return area


def fetch_crossref_abstract(doi_url):
    """Fetch abstract from Crossref API as fallback."""
    doi_id = doi_url.replace("https://doi.org/", "")
    url = CROSSREF_API + doi_id + "?mailto=" + EMAIL

    try:
        if req_lib:
            resp = req_lib.get(url, timeout=15)
        else:
            resp = session.get(url, timeout=15)

        if resp.status_code == 200:
            data = resp.json()
            abstract = data.get("message", {}).get("abstract", "")
            if abstract:
                # Strip HTML tags from abstract
                abstract = re.sub(r"<[^>]+>", "", abstract)
                return abstract
    except Exception as e:
        pass
    return ""


def scrape_article(doi_url):
    """Try INFORMS first, fall back to Crossref abstract."""
    doi_id = doi_url.replace("https://doi.org/", "")
    url = BASE_URL + doi_id

    # ── Attempt 1: INFORMS website ──
    blocked = False
    for attempt in range(5):
        try:
            resp = session.get(url, timeout=30)
            if resp.status_code == 200:
                if "challenge" in resp.text.lower() and len(resp.text) < 5000:
                    if attempt < 4:
                        wait = random.randint(5, 15)
                        print(f"(blocked, retry {attempt+1}/5 in {wait}s)", end=" ", flush=True)
                        time.sleep(wait)
                        continue
                    blocked = True
                    break
                editor, area = extract_editor(resp.text)
                if editor:
                    return editor, area, ""
                # Page loaded but no editor found — try Crossref
                break
            elif resp.status_code == 403:
                if attempt < 4:
                    wait = random.randint(3, 10)
                    print(f"(403, retry {attempt+1}/5 in {wait}s)", end=" ", flush=True)
                    time.sleep(wait)
                    continue
                blocked = True
                break
            else:
                break
        except Exception as e:
            if attempt < 4:
                time.sleep(random.randint(3, 8))
                continue
            blocked = True
            break

    # ── Attempt 2: Crossref API abstract ──
    if blocked:
        print(f"(trying Crossref)", end=" ", flush=True)
    abstract = fetch_crossref_abstract(doi_url)
    if abstract:
        editor, area = extract_editor_from_abstract(abstract)
        if editor:
            return editor, area, "" if not blocked else "(via Crossref)"
    
    if blocked:
        return "", "", "blocked (no Crossref fallback)"
    return "", "", ""


# ── File I/O ──────────────────────────────────────────────────

def find_input_file():
    for name in [OUTPUT_NAME, BASE_NAME]:
        for ext in [".xlsx", ".csv", ".xls"]:
            path = name + ext
            if os.path.exists(path):
                return path
    return None


def read_file(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            return rows, list(reader.fieldnames)
    elif ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        wb.close()
        if not data:
            return [], []
        raw_headers = list(data[0])
        headers = []
        for h in raw_headers:
            if h and str(h).strip():
                headers.append(str(h).strip())
            else:
                break
        num_cols = len(headers)
        rows = []
        for row_data in data[1:]:
            row_dict = {}
            for i in range(num_cols):
                val = row_data[i] if i < len(row_data) else None
                row_dict[headers[i]] = str(val) if val is not None else ""
            rows.append(row_dict)
        return rows, headers
    else:
        print(f"Error: unsupported format '{ext}'")
        sys.exit(1)


def write_file(path, rows, fieldnames):
    import openpyxl
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for c, h in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
        for r, row in enumerate(rows, 2):
            for c, h in enumerate(fieldnames, 1):
                ws.cell(row=r, column=c, value=row.get(h, ""))
        wb.save(path)
        wb.close()
        return path
    except PermissionError:
        for n in range(1, 100):
            alt = path.replace(".xlsx", f"_{n}.xlsx")
            try:
                wb.save(alt)
                wb.close()
                print(f"\n⚠ '{path}' is open. Saved to '{alt}' instead.")
                return alt
            except PermissionError:
                continue
    return None


# ── Fix mode: re-parse badly split editor/area entries ────────

def fix_entries():
    """Find and fix entries where editor field contains 'for the Special' etc."""
    source = find_input_file()
    if not source:
        print("Error: No input file found.")
        sys.exit(1)

    print(f"Reading: {source}")
    rows, fieldnames = read_file(source)

    fixed = 0
    for i, row in enumerate(rows):
        editor = row.get("Accepting Editor", "").strip()
        area = row.get("Area", "").strip()
        if not editor:
            continue

        needs_fix = False
        # Detect bad entries: editor field contains "for the Special" or "served as"
        if "for the special" in editor.lower() or "for the Special" in editor:
            needs_fix = True
        if "served as" in editor.lower():
            needs_fix = True
        # Editor field contains full "accepted by" text
        if "accepted by" in editor.lower():
            needs_fix = True
        # Area contains junk like "Funding:", "Supplemental Material:", etc.
        for junk in ['Funding:', 'Supplemental Material:', 'Conflict of Interest', 'https://doi.org']:
            if junk in area:
                needs_fix = True

        if needs_fix:
            old_editor = editor
            old_area = area
            # Combine editor+area back into raw string and re-parse
            raw = editor
            if area:
                raw = editor + ", " + area
            new_editor, new_area = split_editor_area(raw)
            if new_editor and new_editor != old_editor:
                row["Accepting Editor"] = new_editor
                row["Area"] = new_area
                print(f"  Row {i+1}: \"{old_editor}\" / \"{old_area}\"")
                print(f"       → \"{new_editor}\" / \"{new_area}\"")
                fixed += 1

    if fixed == 0:
        print("\nNo entries need fixing.")
        return

    print(f"\n✓ Fixed {fixed} entries.")
    out_path = OUTPUT_NAME + ".xlsx"
    saved = write_file(out_path, rows, fieldnames)
    print(f"Saved to: {saved}")


# ── Main ──────────────────────────────────────────────────────

def main():
    source = find_input_file()
    if not source:
        print("Error: No input file found.")
        print(f"Place '{BASE_NAME}.csv' or '{BASE_NAME}.xlsx' in this folder.")
        sys.exit(1)

    print(f"Reading: {source}")
    rows, fieldnames = read_file(source)

    if "Accepting Editor" not in fieldnames:
        fieldnames.append("Accepting Editor")
    if "Area" not in fieldnames:
        fieldnames.append("Area")

    for row in rows:
        if "Accepting Editor" not in row:
            row["Accepting Editor"] = ""
        if "Area" not in row:
            row["Area"] = ""

    total = len(rows)
    already_filled = sum(1 for r in rows if r.get("Accepting Editor", "").strip())
    print(f"Loaded {total} rows ({already_filled} already have editor info)")
    print()
    print("Choose mode:")
    print("  1. Scrape editors (normal)")
    print("  2. Fix bad entries (re-parse 'for the Special Issue' etc.)")
    choice = input("Enter 1 or 2 (default 1): ").strip()

    if choice == "2":
        fix_entries()
        return

    start_input = input(f"Start from row (1-{total}, default 1): ").strip()
    start_row = int(start_input) if start_input else 1
    end_input = input(f"End at row (1-{total}, default {total}): ").strip()
    end_row = int(end_input) if end_input else total
    start_row = max(1, min(start_row, total))
    end_row = max(start_row, min(end_row, total))

    print(f"\nProcessing rows {start_row} to {end_row} ({end_row - start_row + 1} rows)")
    print("Ctrl+C once = skip row | Ctrl+C twice = stop & save")
    print("=" * 60)

    filled = 0
    skipped = 0
    failed = 0
    title_skipped = 0
    crossref_filled = 0
    errors = []
    stopped_at = None

    try:
        for i in range(start_row - 1, end_row):
            row = rows[i]
            row_num = i + 1
            doi = row.get("DOI", "").strip()
            status = row.get("Status", "").strip()
            existing_editor = row.get("Accepting Editor", "").strip()
            title = row.get("Title", "").strip()

            # Skip if already filled, no DOI, or status is Other
            if existing_editor or not doi or status == "Other":
                skipped += 1
                continue

            # Skip non-research content by title
            if should_skip_title(title):
                title_skipped += 1
                continue

            title_short = title[:50]
            print(f"[Row {row_num}/{end_row}] {title_short}...", end=" ", flush=True)

            try:
                editor, area, err = scrape_article(doi)
            except KeyboardInterrupt:
                print(" ⏭ SKIPPED (Ctrl+C)")
                skipped += 1
                time.sleep(1)
                continue

            if err and "Crossref" in err:
                # Found via Crossref fallback
                row["Accepting Editor"] = editor
                row["Area"] = area
                print(f"→ {editor}, {area} (Crossref)")
                filled += 1
                crossref_filled += 1
            elif err and not editor:
                print(f"✗ {err}")
                errors.append(f"Row {row_num}: {doi} — {err}")
                failed += 1
            elif editor:
                row["Accepting Editor"] = editor
                row["Area"] = area
                print(f"→ {editor}, {area}")
                filled += 1
            else:
                print("→ (no editor found)")
                failed += 1

            time.sleep(2)

    except KeyboardInterrupt:
        stopped_at = row_num
        print(f"\n\n⛔ STOPPED at row {row_num}. Saving progress...")

    out_path = OUTPUT_NAME + ".xlsx"
    saved = write_file(out_path, rows, fieldnames)

    print()
    print("=" * 60)
    print(f"Results for rows {start_row}-{end_row}:")
    print(f"  Filled:         {filled} ({crossref_filled} via Crossref fallback)")
    print(f"  Skipped:        {skipped} (already filled, Other, or no DOI)")
    print(f"  Title-skipped:  {title_skipped} (non-research content)")
    print(f"  Failed:         {failed}")
    print(f"\nOutput saved to: {saved}")

    if errors:
        print(f"\nErrors ({len(errors)}):")
        for e in errors[:10]:
            print(f"  {e}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more")

    if stopped_at:
        print(f"\n⛔ Stopped early. To resume: run again with start row = {stopped_at}")
    else:
        print(f"\nTo continue: run again with start row = {end_row + 1}")


if __name__ == "__main__":
    main()